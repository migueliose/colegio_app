# docentes/views/estudiantes.py
from django.shortcuts import render, get_object_or_404
from django.views.generic import ListView, TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Q
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from estudiantes.models import Estudiante, Matricula, Acudiente
from gestioncolegio.models import AñoLectivo
from matricula.models import AsignaturaGradoAñoLectivo
from usuarios.models import Docente
from docentes.mixins import DocenteRequiredMixin
from .comportamiento import get_docente_usuario

class ListaEstudiantesView(LoginRequiredMixin, DocenteRequiredMixin, ListView):
    """Vista para listar estudiantes a cargo del docente"""
    model = Estudiante
    template_name = 'docentes/estudiantes_list.html'
    context_object_name = 'estudiantes'
    paginate_by = 20

    def get_docente(self):
        return get_docente_usuario(self.request.user)

    def get_queryset(self):
        docente = self.get_docente()
        if not docente:
            return Estudiante.objects.none()
        
        año_actual = AñoLectivo.objects.filter(estado=True).first()
        if not año_actual:
            return Estudiante.objects.none()
        
        # Obtener grados asignados al docente
        grados_docente = AsignaturaGradoAñoLectivo.objects.filter(
            docente=docente,
            grado_año_lectivo__año_lectivo=año_actual
        ).values_list('grado_año_lectivo__grado_id', flat=True).distinct()
        
        if not grados_docente:
            return Estudiante.objects.none()
        
        # Obtener estudiantes matriculados en esos grados
        estudiantes = Estudiante.objects.filter(
            matriculas__grado_año_lectivo__grado_id__in=grados_docente,
            matriculas__año_lectivo=año_actual,
            matriculas__estado__in=['ACT', 'PEN']
        ).select_related(
            'usuario'
        ).prefetch_related(
            'acudientes__acudiente'
        ).distinct()
        
        # Aplicar filtros
        grado_id = self.request.GET.get('grado')
        search = self.request.GET.get('search', '').strip()
        
        if grado_id:
            estudiantes = estudiantes.filter(
                matriculas__grado_año_lectivo__grado_id=grado_id
            )
        
        if search:
            estudiantes = estudiantes.filter(
                Q(usuario__nombres__icontains=search) |
                Q(usuario__apellidos__icontains=search) |
                Q(usuario__numero_documento__icontains=search) |
                Q(usuario__email__icontains=search)
            )
        
        return estudiantes.order_by('usuario__apellidos', 'usuario__nombres')
    
    def get_grados_docente(self, docente):
        """Obtener grados asignados al docente"""
        año_actual = AñoLectivo.objects.filter(estado=True).first()
        if not año_actual:
            return []
        
        from academico.models import Grado
        
        grados_ids = AsignaturaGradoAñoLectivo.objects.filter(
            docente=docente,
            grado_año_lectivo__año_lectivo=año_actual
        ).values_list('grado_año_lectivo__grado_id', flat=True).distinct()
        
        return Grado.objects.filter(id__in=grados_ids).order_by('nombre')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        docente = self.get_docente()
        año_actual = AñoLectivo.objects.filter(estado=True).first()
        
        if not docente or not año_actual:
            context['grados'] = []
            return context
        
        # Obtener grados del docente
        context['grados'] = self.get_grados_docente(docente)
        
        # Mantener filtros actuales
        context['filtro_grado'] = self.request.GET.get('grado', '')
        context['filtro_search'] = self.request.GET.get('search', '')
        
        # Estadísticas
        estudiantes = self.get_queryset()
        context['total_estudiantes'] = estudiantes.count()
        
        # Obtener distribución por grado
        distribucion_grados = {}
        for estudiante in estudiantes:
            matricula_actual = estudiante.matricula_actual
            if matricula_actual and matricula_actual.grado_año_lectivo:
                grado_nombre = matricula_actual.grado_año_lectivo.grado.nombre
                distribucion_grados[grado_nombre] = distribucion_grados.get(grado_nombre, 0) + 1
        
        context['distribucion_grados'] = distribucion_grados
        
        return context


class ExportarEstudiantesExcelView(LoginRequiredMixin, DocenteRequiredMixin, TemplateView):
    """Vista para exportar estudiantes a Excel"""
    
    def get_docente(self):
        return get_docente_usuario(self.request.user)
    
    def get(self, request, *args, **kwargs):
        docente = self.get_docente()
        if not docente:
            return HttpResponse("No tiene permisos", status=403)
        
        año_actual = AñoLectivo.objects.filter(estado=True).first()
        if not año_actual:
            return HttpResponse("No hay año lectivo activo", status=400)
        
        # Obtener parámetros de filtro
        grado_id = request.GET.get('grado')
        
        # Obtener grados asignados al docente
        grados_docente = AsignaturaGradoAñoLectivo.objects.filter(
            docente=docente,
            grado_año_lectivo__año_lectivo=año_actual
        ).values_list('grado_año_lectivo__grado_id', flat=True).distinct()
        
        # Filtrar estudiantes
        estudiantes = Estudiante.objects.filter(
            matriculas__grado_año_lectivo__grado_id__in=grados_docente,
            matriculas__año_lectivo=año_actual,
            matriculas__estado__in=['ACT', 'PEN']
        ).select_related(
            'usuario'
        ).prefetch_related(
            'acudientes__acudiente'
        ).distinct()
        
        if grado_id:
            estudiantes = estudiantes.filter(
                matriculas__grado_año_lectivo__grado_id=grado_id
            )
        
        estudiantes = estudiantes.order_by('usuario__apellidos', 'usuario__nombres')
        
        # Crear libro de Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Estudiantes"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        cell_fill_even = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell_fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Encabezados
        headers = [
            'N°', 'Documento', 'Nombres', 'Apellidos', 'Fecha Nacimiento', 
            'Edad', 'Sexo', 'Grado', 'Email', 'Teléfono', 'Dirección',
            'Acudiente 1', 'Teléfono Acudiente 1', 'Parentesco 1',
            'Acudiente 2', 'Teléfono Acudiente 2', 'Parentesco 2',
            'Estado Matrícula'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Datos
        row_num = 2
        for idx, estudiante in enumerate(estudiantes, 1):
            usuario = estudiante.usuario
            matricula_actual = estudiante.matricula_actual
            
            # Obtener acudientes
            acudientes = estudiante.acudientes.all()
            acudiente1 = acudientes[0] if acudientes else None
            acudiente2 = acudientes[1] if len(acudientes) > 1 else None
            
            # Calcular edad
            edad = None
            if usuario.fecha_nacimiento:
                today = timezone.now().date()
                edad = today.year - usuario.fecha_nacimiento.year - (
                    (today.month, today.day) < (usuario.fecha_nacimiento.month, usuario.fecha_nacimiento.day)
                )
            
            # Preparar datos
            data = [
                idx,  # N°
                usuario.numero_documento or '',
                usuario.nombres or '',
                usuario.apellidos or '',
                usuario.fecha_nacimiento.strftime('%d/%m/%Y') if usuario.fecha_nacimiento else '',
                edad or '',
                usuario.get_sexo_display() if usuario.sexo else '',
                matricula_actual.grado_año_lectivo.grado.nombre if matricula_actual else '',
                usuario.email or '',
                usuario.telefono or '',
                usuario.direccion or '',
                acudiente1.acudiente.get_full_name() if acudiente1 else '',
                acudiente1.acudiente.telefono if acudiente1 and hasattr(acudiente1.acudiente, 'telefono') else '',
                acudiente1.parentesco if acudiente1 else '',
                acudiente2.acudiente.get_full_name() if acudiente2 else '',
                acudiente2.acudiente.telefono if acudiente2 and hasattr(acudiente2.acudiente, 'telefono') else '',
                acudiente2.parentesco if acudiente2 else '',
                matricula_actual.get_estado_display() if matricula_actual else ''
            ]
            
            # Escribir fila
            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
                
                # Alternar colores de fila
                if row_num % 2 == 0:
                    cell.fill = cell_fill_even
                else:
                    cell.fill = cell_fill_odd
                
                # Alineación
                if col_num in [1, 2, 6, 7, 12, 13, 15, 16, 18]:  # Columnas numéricas o cortas
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.alignment = Alignment(horizontal="left")
            
            row_num += 1
        
        # Ajustar anchos de columna
        column_widths = {
            'A': 5,   # N°
            'B': 15,  # Documento
            'C': 20,  # Nombres
            'D': 20,  # Apellidos
            'E': 15,  # Fecha Nacimiento
            'F': 8,   # Edad
            'G': 10,  # Sexo
            'H': 15,  # Grado
            'I': 25,  # Email
            'J': 15,  # Teléfono
            'K': 30,  # Dirección
            'L': 25,  # Acudiente 1
            'M': 15,  # Teléfono Acudiente 1
            'N': 15,  # Parentesco 1
            'O': 25,  # Acudiente 2
            'P': 15,  # Teléfono Acudiente 2
            'Q': 15,  # Parentesco 2
            'R': 15,  # Estado Matrícula
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Agregar información del reporte
        ws.cell(row=row_num + 1, column=1, value="Reporte generado el:").font = Font(bold=True)
        ws.cell(row=row_num + 1, column=2, value=timezone.now().strftime('%d/%m/%Y %H:%M'))
        
        ws.cell(row=row_num + 2, column=1, value="Docente:").font = Font(bold=True)
        ws.cell(row=row_num + 2, column=2, value=docente.usuario.get_full_name())
        
        ws.cell(row=row_num + 3, column=1, value="Año lectivo:").font = Font(bold=True)
        ws.cell(row=row_num + 3, column=2, value=año_actual.anho)
        
        ws.cell(row=row_num + 4, column=1, value="Total estudiantes:").font = Font(bold=True)
        ws.cell(row=row_num + 4, column=2, value=estudiantes.count())
        
        # Preparar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"estudiantes_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Guardar libro
        wb.save(response)
        
        return response


class DetalleEstudianteView(LoginRequiredMixin, DocenteRequiredMixin, TemplateView):
    """Vista para ver detalles de un estudiante específico"""
    template_name = 'docentes/estudiante_detalle.html'
    
    def get_docente(self):
        return get_docente_usuario(self.request.user)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        docente = self.get_docente()
        estudiante_id = self.kwargs.get('pk')
        
        if not docente:
            return context
        
        # Obtener estudiante
        try:
            estudiante = Estudiante.objects.select_related(
                'usuario'
            ).prefetch_related(
                'acudientes__acudiente',
                'matriculas__grado_año_lectivo__grado',
                'matriculas__sede'
            ).get(id=estudiante_id)
        except Estudiante.DoesNotExist:
            return context
        
        # Verificar que el estudiante esté en los grados del docente
        año_actual = AñoLectivo.objects.filter(estado=True).first()
        if año_actual:
            grados_docente = AsignaturaGradoAñoLectivo.objects.filter(
                docente=docente,
                grado_año_lectivo__año_lectivo=año_actual
            ).values_list('grado_año_lectivo__grado_id', flat=True).distinct()
            
            estudiante_valido = estudiante.matriculas.filter(
                grado_año_lectivo__grado_id__in=grados_docente,
                año_lectivo=año_actual
            ).exists()
            
            if not estudiante_valido:
                return context
        
        # Obtener información académica
        from comportamiento.models import Comportamiento, Asistencia
        from estudiantes.models import Nota
        from matricula.models import PeriodoAcademico
        
        periodo_actual = PeriodoAcademico.objects.filter(
            año_lectivo=año_actual,
            estado=True
        ).first()
        
        # Comportamientos recientes
        comportamientos = Comportamiento.objects.filter(
            estudiante=estudiante
        ).select_related(
            'periodo_academico__periodo',
            'docente__usuario'
        ).order_by('-fecha')[:10]
        
        # Asistencia
        asistencia_stats = {'porcentaje': 0, 'total': 0, 'asistencias': 0}
        if periodo_actual:
            asistencias = Asistencia.objects.filter(
                estudiante=estudiante,
                periodo_academico=periodo_actual
            )
            total = asistencias.count()
            if total > 0:
                asistencia_stats = {
                    'porcentaje': round(asistencias.filter(estado='A').count() / total * 100, 1),
                    'total': total,
                    'asistencias': asistencias.filter(estado='A').count()
                }
        
        # Notas recientes
        notas = Nota.objects.filter(
            estudiante=estudiante
        ).select_related(
            'asignatura_grado_año_lectivo__asignatura',
            'periodo_academico__periodo'
        ).order_by('-periodo_academico__fecha_inicio')[:10]
        
        # Historial académico
        historial_matriculas = estudiante.matriculas.select_related(
            'grado_año_lectivo__grado',
            'sede',
            'año_lectivo'
        ).order_by('-año_lectivo__anho')
        
        context.update({
            'estudiante': estudiante,
            'comportamientos': comportamientos,
            'asistencia_stats': asistencia_stats,
            'notas': notas,
            'historial_matriculas': historial_matriculas,
            'periodo_actual': periodo_actual,
        })
        
        return context